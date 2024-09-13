""" Функции по чтению осей с файла и их преобразованию """

import csv
import re
from typing import NoReturn

import lasio
import numpy as np
# import pandas as pd
from openpyxl import load_workbook
from xls2xlsx import XLS2XLSX
from excel_parcer.models import *
from openpyxl.utils import get_column_letter, column_index_from_string

from ..models import List, Device, Data
from ..serializer import DataSerializer
from datetime import datetime

def toFixed(numObj, digits=0):
    return f"{numObj:.{digits}f}"


def parcing_manually(path, manually_depth, manually_gx, manually_gy, manually_gz, manually_bx, manually_by, manually_bz,
                     manually_import=None) -> tuple:
    """
    Считываем данные с экселя осей по указанным параметрам. Возращает list с list'ами.
    Лист замеров - лист нижней абстракции это набор параметров глубина - оси.
    """
    try:
        manually_import = int(manually_import)
    except ValueError:
        manually_import = None
    data: list = [[], [], [], [], [], [], []]

    if "las" in re.findall(r".(las)", path):
        """Для обработки las"""
        las_file = lasio.read(path)

        for i, name in enumerate(
                [manually_bz, manually_by, manually_bx, manually_gz, manually_gy, manually_gx, manually_depth]):
            if name in las_file.keys():
                data[i] = np.asarray(las_file[name])
        data = np.asarray(data)
        data = data[:, ~np.isnan(data).any(axis=0)]
    # FIXME вынести все условия в функции
    elif "csv" in re.findall(r".(csv)", path):
        """Для обработки csv"""
        with open(path, 'r', newline='') as csvfile:
            i: int = 0
            index_List = [manually_bz, manually_by, manually_bx, manually_gz, manually_gy, manually_gx, manually_depth]
            datareader = csv.reader(csvfile, delimiter=';', quotechar='|')
            if manually_import is None:
                manually_import = 0
            for raw in datareader:
                print(raw)
                i += 1
                if len(raw) == 0 or i < manually_import:
                    continue
                for d_index, n_index in enumerate(index_List):
                    try:
                        data[d_index].append(float(raw[int(n_index) - 1]))
                    except ValueError:
                        pass
    elif "txt" in re.findall(r".(txt)", path):
        """Для обработки txt"""
        index_List = [manually_bz, manually_by, manually_bx, manually_gz, manually_gy, manually_gx, manually_depth]
        start_index = (manually_import if manually_import is not None else 31)
        with open(path, encoding="utf-8") as f:
            for i in range(int(start_index) - 1):
                f.readline()  # считываем ненужные строки
            for raw in f.readlines():
                try:
                    raw_list = raw.replace('\n', '').replace('\r', '').replace('\t\t', '\t').replace('\t', ',').split(
                        sep=',')
                    # print(raw_list)
                    for index in index_List:  # проверка считанного замера
                        float(raw_list[int(index) - 1])
                    for d_index, n_index in enumerate(index_List):
                        data[d_index].append(float(raw_list[int(n_index) - 1]))
                except Exception as e:
                    print('Пропуск замера. Ошибка:\n', e)
                    continue
    elif 'sur' in re.findall(r".(sur)", path):
        index_List = [manually_bz, manually_by, manually_bx, manually_gz, manually_gy, manually_gx, manually_depth]
        start_index = (manually_import if manually_import is not None else 2)
        with open(path, encoding="utf-8") as f:
            for i in range(int(start_index)):
                f.readline()
            for raw in f.readlines():
                raw_list = raw.replace('\n', '').split()
                if len(raw_list) < 7:  # в строке отсутсвует разделитель
                    continue
                for d_index, n_index in enumerate(index_List):
                    data[d_index].append(float(raw_list[int(n_index) - 1]))
    else:
        """Для обработки excel"""
        try:

            wb = load_workbook(filename=path, data_only=True)
            #wb = pd.read_excel(path, engine='openpyxl')
            print('why')
        except:

            x2x = XLS2XLSX(path)
            print('xlsx', x2x)
            wb = x2x.to_xlsx()

        sheet = wb.active
        print('sheet', sheet)
        for i, name in enumerate(
                [manually_bz, manually_by, manually_bx, manually_gz, manually_gy, manually_gx, manually_depth]):
            print('i, name', i, name)
            colum_id = column_index_from_string(name)

            for row_id in range(int(manually_import), sheet.max_row + 1):
                cell = sheet.cell(row_id, colum_id)
                if cell.value is not None:
                    data[i].append(cell.value)
                else:
                    data[i].append('')

    return tuple(zip(*data[::-1]))


def new_parcing(path):
    try:
        wb = load_workbook(filename=path)
    except:
        x2x = XLS2XLSX(path)
        wb = x2x.to_xlsx()
    sheet = wb.worksheets[0]
    headers = []
    measures = []
    depth = []
    data = []
    GX = []
    GY = []
    GZ = []
    BX = []
    BY = []
    BZ = []
    unique_1 = 0
    unique_2 = 0
    unique_3 = 0
    unique_4 = 0
    unique_5 = 0
    unique_6 = 0
    unique_7 = 0

    for i in sheet.columns:
        for j in range(len(i)):
            for k in List.objects.first().depth.split(';'):
                if i[j].value == k:
                    if unique_1 == 0:
                        unique_1 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    depth.append(z.value)

            for k in List.objects.first().CX.split(';'):
                if i[j].value == k:
                    if unique_2 == 0:
                        unique_2 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    GX.append(z.value)

            for k in (List.objects.first().CY).split(';'):
                if (i[j].value == k):
                    if (unique_3 == 0):
                        unique_3 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    GY.append(z.value)

            for k in (List.objects.first().CZ).split(';'):
                if (i[j].value == k):
                    if (unique_4 == 0):
                        unique_4 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    GZ.append(z.value)

            for k in (List.objects.first().BX).split(';'):
                if (i[j].value == k):
                    if (unique_5 == 0):
                        unique_5 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    BX.append(z.value)

            for k in (List.objects.first().BY).split(';'):
                if (i[j].value == k):
                    if (unique_6 == 0):
                        unique_6 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    BY.append(z.value)

            for k in (List.objects.first().BZ).split(';'):
                if (i[j].value == k):
                    if (unique_7 == 0):
                        unique_7 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    BZ.append(z.value)
    for i in range(7):
        measures.append("-")

    data.append(headers)
    data.append(measures)
    for i in range(len(depth)):
        result = []
        result.append(depth[i])
        result.append(GX[i])
        result.append(GY[i])
        result.append(GZ[i])
        result.append(BX[i])
        result.append(BY[i])
        result.append(BZ[i])
        # print(result)
        data.append(result)
    return data


def new_measurements(data, name):
    """
    Тоже бы переписать когда-нибудь. Здесь преобразуем данные по указанным правилам.
    Приводим к одному формату.
    """
    result = []
    for i in range(len(data)):
        new_data = []
        for j in range(len(data[i])):
            if '' in data[i]:
                continue
            if j == 0:  # depth[i]
                new_data.append(toFixed((round(float(eval("((data[i])[j])")), 2)), 2))
            if j == 1:  # GX[i]
                new_data.append(
                    toFixed((round(
                        float(eval("float(((data[i])[j]))" + Device.objects.get(device_title=name).CX)), 7)),
                        7))
            if j == 2:  # GY[i]
                new_data.append(
                    toFixed((round(
                        float(eval("float(((data[i])[j]))" + Device.objects.get(device_title=name).CY)), 7)),
                        7))
            if j == 3:  # GZ[i]
                new_data.append(
                    toFixed((round(
                        float(eval("float(((data[i])[j]))" + Device.objects.get(device_title=name).CZ)), 7)),
                        7))
            if j == 4:  # BX[i]
                new_data.append(
                    toFixed((round(
                        float(eval("float(((data[i])[j]))" + Device.objects.get(device_title=name).BX)), 2)),
                        2))
            if j == 5:  # BY[i]
                new_data.append(
                    toFixed((round(
                        float(eval("float(((data[i])[j]))" + Device.objects.get(device_title=name).BY)), 2)),
                        2))
            if j == 6:  # BZ[i]
                new_data.append(
                    toFixed((round(
                        float(eval("float(((data[i])[j]))" + Device.objects.get(device_title=name).BZ)), 2)),
                        2))
        result.append(new_data)

    return result


# def dbdata(run):
#     data = [["м", "м/с^2", "м/с^2", "м/с^2", "нТ", "нТ", "нТ"]]
#     for i in Data.objects.filter(run=Run.objects.get(run_title=run)):
#         data.append([i.depth, i.CX, i.CY, i.CZ, i.BX, i.BY, i.BZ])
#     return data


def write_to_bd(data: list[list], run: object) -> NoReturn:
    """
    Записываем преобразованные данные в БД, возвращаем conflict словарь старых и новых значений
    которые будут перезаписаны
    """
    create_obj = list()
    conflict = {'old': list(), 'new': list()}
    date=datetime.now()
    for rows in data:
        try:
            if len(rows) < 7:  # проверка на наличие всех осей
                continue
            if float(rows[0]) < 0:  # глубина не может быть отрицательной
                continue
        except Exception as e:
            print(e)
            continue

        try:
            bd_data = Data.objects.get(depth=rows[0], run=run)
            # Если замер не изменился не записываем его
            if bd_data.CX == float(rows[1]) and bd_data.CY == float(rows[2]) and bd_data.CZ == float(rows[3]) and \
                    bd_data.BX == float(rows[4]) and bd_data.BY == float(rows[5]) and bd_data.BZ == float(rows[6]):
                continue
            else:
                conflict['old'].append(DataSerializer(bd_data).data)
                conflict['new'].append(DataSerializer(Data(depth=rows[0],
                                                           CX=rows[1],
                                                           CY=rows[2],
                                                           CZ=rows[3],
                                                           BX=rows[4],
                                                           BY=rows[5],
                                                           BZ=rows[6],
                                                           date=date
                                                           )).data)

        except Data.DoesNotExist:
            create_obj.append(Data(
                depth=rows[0],
                run=run,
                CX=rows[1],
                CY=rows[2],
                CZ=rows[3],
                BX=rows[4],
                BY=rows[5],
                BZ=rows[6],
                date=date,
                in_statistics=True, ))

    Data.objects.bulk_create(create_obj)
    current_date=0
    if len(create_obj)>0:
        current_date=date
    return conflict, current_date


def convert_sign(eval_str: str) -> str:
    """ Меняет знаки в выражениях для пересчёта """
    if eval_str.find('*') != -1:
        return eval_str.replace('*', '/')
    elif eval_str.find('/') != -1:
        return eval_str.replace('/', '*')
    elif eval_str.find('+') != -1:
        return eval_str.replace('+', '-')
    elif eval_str.find('-') != -1:
        return eval_str.replace('-', '+')
