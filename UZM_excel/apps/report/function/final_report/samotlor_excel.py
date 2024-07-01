import io
import os
import openpyxl
from math import sqrt
from typing import NoReturn

from matplotlib import pyplot as plt
from openpyxl.drawing.image import Image
from Field.models import Run
from report.function.graffic import get_graphics
from report.function.model_service import get_data, get_waste
from excel_parcer.models import Data


def final_report_excel(my_run: object) -> str:
    """
    Excel файл по Самотлору (приложение к финальному отчёту)
    """
    template = os.getcwd() + "\\files\\Шаблон\\Самотлорское_Приложение.xlsx"
    excel_file = openpyxl.load_workbook(template)

    my_wellbore = my_run.section.wellbore
    runs = Run.objects.filter(section__wellbore=my_wellbore)
    all_data = get_data(runs)

    # построение графика,  other_data - тут лежат TVD, угол, азимут
    other_data, waste_word = get_graphics(all_data, my_wellbore)
    # добавляем график проекции
    grafic = Image(os.getcwd() + f"\\files\\Report_out\\{my_wellbore}.png")
    excel_file['Проекции'].add_image(grafic, 'A1')

    # если бурим по траектории ИГиРГИ заменяем замеры ннб на план
    nnb_data = all_data['Статические замеры ННБ'] if not my_wellbore.igirgi_drilling else \
        all_data['Плановая траектория интерп']
    write_data(excel_file, nnb_data, all_data['Статические замеры ИГИРГИ'], other_data, my_run)
    samotlor_QualityCharts(excel_file, my_wellbore)

    excel_file.save(os.getcwd() + f"\\files\\Report_out\\{my_wellbore}_Приложение")
    return os.getcwd() + f"\\files\\Report_out\\{my_wellbore}_Приложение"


def samotlor_QualityCharts(excel, wellbore_obj):
    """Заполняет информацией слайдов контроля качества"""
    # !!! Данные для графика !!!
    runs = Run.objects.filter(section__wellbore=wellbore_obj)

    context = {'depth': list(),
               'depthGtotal': list(),
               'depthGref': list(),
               'depthGmax': list(),
               'depthGmin': list(),
               'depthDipraw': list(),
               'depthDipref': list(),
               'depthDipmax': list(),
               'depthDipmin': list(),
               'depthDipcorr': list(),
               "depthBoxy": list(),
               "depthBz": list(),
               "depthGoxy": list(),
               "depthGz": list(),
               "depthBtotal": list(),
               "depthBref": list(),
               "depthBmax": list(),
               "depthBmin": list(),
               "depthBcorr": list(),
               }
    for run in runs:
        surveys = Data.objects.filter(run=run).order_by('depth')
        for survey in surveys:
            context['depth'].append(survey.depth)
            # График Gtotal
            context['depthGtotal'].append(survey.Gtotal())
            context['depthGref'].append(wellbore_obj.well_name.gtotal)
            context['depthGmax'].append(wellbore_obj.well_name.max_gtotal())
            context['depthGmin'].append(wellbore_obj.well_name.min_gtotal())
            # График Goxy-Gz
            context['depthGoxy'].append(survey.get_goxy())
            context['depthGz'].append(survey.CZ)
            # График Btotal
            context['depthBtotal'].append(survey.Btotal())
            context['depthBref'].append(wellbore_obj.well_name.btotal)
            context['depthBmax'].append(wellbore_obj.well_name.max_btotal())
            context['depthBmin'].append(wellbore_obj.well_name.min_btotal())
            context['depthBcorr'].append(survey.Btotal_corr)
            # График Dip
            context['depthDipraw'].append(survey.Dip())
            context['depthDipref'].append(wellbore_obj.well_name.dip)
            context['depthDipmax'].append(wellbore_obj.well_name.max_dip())
            context['depthDipmin'].append(wellbore_obj.well_name.min_dip())
            context['depthDipcorr'].append(survey.DIP_corr)

    # отступы графика от краёв
    margins = {  # +++
        "left": 0.09,
        "bottom": 0.2,
        "right": 0.85,
        "top": 0.90
    }
    bbox = (1.02, 1)

    # отступы легенды графика от угла
    # !!! Графики контроля качества 1/2 !!!
    image_stream = io.BytesIO()
    plt.figure(figsize=(12, 4))
    plt.subplots_adjust(**margins)
    plt.plot(context['depth'], context['depthGtotal'], color='#D6D64C', label='Gtotal')
    plt.plot(context['depth'], context['depthGref'], color='g', label='Gref')
    plt.plot(context['depth'], context['depthGmax'], '--r')
    plt.plot(context['depth'], context['depthGmin'], '--r')
    plt.title('График напряженности гравитационного поля', fontsize=12)
    plt.xlabel('Измеренная глубина, м', color='gray', fontsize=10)
    plt.ylabel('Gtotal, G', color='gray', fontsize=10)
    plt.legend(bbox_to_anchor=bbox, loc='upper left', fontsize=10)
    plt.grid(True)
    plt.savefig(image_stream)
    grafic = Image(image_stream)
    excel['Графики контроля качества'].add_image(grafic, 'A1')

    # График магнитного наклонения
    image_stream = io.BytesIO()
    plt.figure(figsize=(12, 4))
    plt.subplots_adjust(**margins)
    plt.plot(context['depth'], context['depthDipraw'], color='#D6D64C', label='DipRaw')
    plt.plot(context['depth'], context['depthDipref'], color='green', label='DipRef')
    plt.plot(context['depth'], context['depthDipmax'], '--r')
    plt.plot(context['depth'], context['depthDipmin'], '--r')
    plt.plot(context['depth'], context['depthDipcorr'], color='blue', label='DipCorr')
    plt.title('График магнитного наклонения', fontsize=12)
    plt.xlabel('Измеренная глубина, м', color='gray', fontsize=10)
    plt.ylabel('Магнитное наклонение, гр', color='gray', fontsize=10)
    plt.legend(bbox_to_anchor=bbox, loc='upper left', fontsize=10)
    plt.grid(True)
    plt.savefig(image_stream)
    grafic = Image(image_stream)
    excel['Графики контроля качества'].add_image(grafic, 'A41')

    # График напряженности геомагнитного поля
    image_stream = io.BytesIO()
    plt.figure(figsize=(12, 4))
    plt.subplots_adjust(**margins)
    plt.plot(context['depth'], context['depthBtotal'], color='#D6D64C', label='Bt_RAW')
    plt.plot(context['depth'], context['depthBcorr'], color='blue', label='Bcorr')
    plt.plot(context['depth'], context['depthBmax'], '--r')
    plt.plot(context['depth'], context['depthBref'], color='green', label='Bref')
    plt.plot(context['depth'], context['depthBmin'], '--r')
    plt.title('График напряженности геомагнитного поля', fontsize=12)
    plt.xlabel('Измеренная глубина, м', color='gray', fontsize=10)
    plt.ylabel('Напряженность геомагнитного поля, нТл', color='gray', fontsize=6)
    plt.legend(bbox_to_anchor=bbox, loc='upper left', fontsize=10)
    plt.grid(True)
    plt.savefig(image_stream)
    grafic = Image(image_stream)
    excel['Графики контроля качества'].add_image(grafic, 'A21')


def write_data(excel: openpyxl.workbook.workbook.Workbook,
               nnb: dict, igirgi: dict, other: dict,
               Run: object):
    """
    Первая страница запись (Данные)
    Возвращаем последние посчитанные отходы для наименования файла
    """
    well_obj = Run.section.wellbore.well_name
    write_hat(excel['Данные'], well_obj)  # шапка для страницы
    samotlor_hat(excel['Данные'], well_obj)
    waste = samotlor_data(excel['Данные'], nnb, igirgi, other, well_obj)
    return waste


def samotlor_data(excel_sheet: openpyxl.workbook.workbook.Workbook,
                  nnb: dict, igirgi: dict, other: dict,
                  my_well: object) -> tuple[float, float, float]:
    """Заполняем таблицу значениями"""
    row = 18
    count = 0  # первую строку не выводим
    # запись в ячейки
    for meas in zip(igirgi['Глубина'], igirgi['Угол'], igirgi['Азимут'], nnb['Угол'], nnb['Азимут'],
                    other['igirgi_TVDSS'], other['igirgi_delta_x'], other['igirgi_delta_y'], other['igirgi_TVD'],
                    other['nnb_delta_x'], other['nnb_delta_y'], other['nnb_TVD'], igirgi['Комментарий'],
                    igirgi['Рейс']):
        if count == 0:
            count += 1
            continue
        excel_sheet.cell(row=row, column=1).value = (meas[13] if meas[13] != -1 else 'материнский ствол')  # Номер рейса
        excel_sheet.cell(row=row, column=2).value = round(meas[0], 2)  # Глубина
        excel_sheet.cell(row=row, column=3).value = round(meas[1], 2)  # Зенитный угол
        excel_sheet.cell(row=row, column=4).value = round(meas[2], 2)  # Азимут

        if my_well.dec is None:  # Азимут магнитный
            excel_sheet.cell(row=row, column=5).value = '-'
        else:
            azimut = round(meas[2] - my_well.total_correction, 2)
            excel_sheet.cell(row=row, column=5).value = azimut if azimut > 0 else azimut + 360

        excel_sheet.cell(row=row, column=6).value = round(meas[3], 2)  # ННБ зенитный угол
        excel_sheet.cell(row=row, column=7).value = round(meas[4], 2)  # ННБ азимут
        excel_sheet.cell(row=row, column=8).value = round(meas[1] - meas[3], 2)  # разница зенитный угол
        dif_Az = meas[2] - meas[4]  # разница азимут
        dif_Az = dif_Az if dif_Az <= 300 else dif_Az - 360
        dif_Az = dif_Az if dif_Az >= -300 else dif_Az + 360
        excel_sheet.cell(row=row, column=9).value = round(dif_Az, 2)

        # пошли отходы
        Ex = (my_well.EX if my_well.EX is not None else 0)
        Ny = (my_well.NY if my_well.NY is not None else 0)
        X_nnb = Ex + meas[10]
        Y_nnb = Ny + meas[9]

        X_igirgi = Ex + meas[7]
        Y_igigri = Ny + meas[6]

        excel_sheet.cell(row=row, column=10).value = round(  # отход по горизонтали
            sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2), 2)
        excel_sheet.cell(row=row, column=11).value = round(meas[11] - meas[8], 2)  # отход по вертикали
        excel_sheet.cell(row=row, column=12).value = round(  # отход общий
            sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2 +
                 (meas[11] - meas[8]) ** 2), 2)

        excel_sheet.cell(row=row, column=13).value = meas[12]
        if row == (17 + len(igirgi['Глубина']) - 1):  # забираем последние значения отходов
            return round(sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2), 2), round(meas[11] - meas[8], 2), \
                round(sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2 + (meas[11] - meas[8]) ** 2), 2)
        row += 1


def samotlor_hat(excel_sheet: openpyxl.workbook.workbook.Workbook.worksheets, my_well: object) -> NoReturn:
    """Шапка колонок таблицы для самотлорского отчёта"""
    excel_sheet['D16'].value = '-' if my_well.north_direction is None else my_well.get_north_direction()
    excel_sheet['G16'].value = '-' if my_well.north_direction is None else my_well.get_north_direction()
    excel_sheet['I16'].value = '-' if my_well.north_direction is None else my_well.get_north_direction()
    if my_well.wellbores.all()[0].igirgi_drilling:
        excel_sheet['F14'].value = 'Плановая траектория'


def write_hat(excel_sheet: openpyxl.workbook.workbook.Workbook.worksheets, my_well) -> NoReturn:
    """
    Записываем шапку Excel страницы по данным модели Well
    """
    excel_sheet['C3'].value = my_well.pad_name.field.field_name
    excel_sheet['C4'].value = my_well.pad_name.pad_name
    excel_sheet['C5'].value = my_well.well_name
    excel_sheet['C6'].value = '-' if my_well.RKB is None else my_well.RKB
    excel_sheet['C7'].value = '-' if my_well.coordinate_system is None else my_well.coordinate_system
    excel_sheet['C8'].value = '-' if my_well.latitude is None else my_well.latitude
    excel_sheet['C9'].value = '-' if my_well.longtitude is None else my_well.longtitude
    excel_sheet['C10'].value = '-' if my_well.NY is None else my_well.NY
    excel_sheet['C11'].value = '-' if my_well.EX is None else my_well.EX
    excel_sheet['H3'].value = '-' if my_well.geomagnetic_model is None else my_well.geomagnetic_model
    excel_sheet['H4'].value = '-' if my_well.geomagnetic_date is None else my_well.geomagnetic_date.strftime('%d-%m-%Y')
    excel_sheet['H5'].value = '-' if my_well.north_direction is None else my_well.get_north_direction()
    excel_sheet['H6'].value = '-' if my_well.btotal is None else my_well.btotal
    excel_sheet['H7'].value = '-' if my_well.dip is None else my_well.dip
    excel_sheet['H8'].value = '-' if my_well.dec is None else my_well.dec
    excel_sheet['H9'].value = '-' if my_well.grid_convergence is None else my_well.grid_convergence
    excel_sheet['H10'].value = '-' if my_well.total_correction is None else my_well.total_correction
    excel_sheet['H11'].value = '-' if my_well.gtotal is None else my_well.gtotal

